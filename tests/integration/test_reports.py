from pathlib import Path

from openpyxl import load_workbook

from calculadora_trabalhista.calculation.engine import CalculationEngine
from calculadora_trabalhista.ingestion.json_provider import JsonProcessFactsProvider
from calculadora_trabalhista.reports.exporters import export_json, export_markdown, export_xlsx


def test_exports_are_readable(tmp_path: Path) -> None:
    fixture = Path(__file__).parents[1] / "fixtures" / "process_facts_synthetic.json"
    result = CalculationEngine().calculate(JsonProcessFactsProvider(fixture).load())
    json_path = export_json(result, tmp_path / "result.json")
    md_path = export_markdown(result, tmp_path / "Tabela_Pedidos.md")
    table_path, memory_path = export_xlsx(result, tmp_path)
    assert json_path.exists() and md_path.read_text(encoding="utf-8").startswith("# Tabela")
    assert load_workbook(table_path)["Tabela_Pedidos"].max_row > 2
    assert load_workbook(memory_path)["Memoria_Calculo"].max_row > 2
